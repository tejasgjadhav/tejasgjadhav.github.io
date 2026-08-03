"""Import the GlobalAAR monthly Excel MIS workbook (MASTERS, DATA ENTRY,
REWORK EFFICIENCY) into the SQLite database.

Upsert rule: for each imported shift row, any existing rows with the same
(date, machine, shift, part, operator) are replaced. All derived numbers are
recomputed with calc.py; rows whose recomputed OEE/OK/rejection differ from the
workbook's own stored values are listed in the import report.
"""
import openpyxl

from . import calc
from .models import DEFECT_TYPES, DOWNTIME_REASONS, get_setting

# DATA ENTRY sheet layout (1-based column indexes)
COL_DATE, COL_MACHINE, COL_SHIFT, COL_SUP, COL_OP, COL_PART = 1, 2, 3, 4, 5, 6
COL_PRODUCTION = 12
COL_OK, COL_REJ, COL_OEE = 13, 14, 69
DEFECT_COL_START = 17     # Q .. AK  (21 defect columns)
DOWNTIME_COL_START = 38   # AL .. BG (22 reason columns)
COL_REMARKS, COL_PLANNED = 70, 71


def _cell_date(v):
    if v is None:
        return None
    s = str(v)
    return s[:10] if len(s) >= 10 and s[4] == "-" else None


def _n(v):
    """Numeric cell → float or None."""
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def import_workbook(con, path):
    """Import an .xlsx file. Returns a report dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    report = {"parts_added": 0, "masters_added": 0, "entries": 0, "replaced": 0,
              "rework": 0, "mismatches": [], "skipped": []}
    if "MASTERS" in wb.sheetnames:
        _import_masters(con, wb["MASTERS"], report)
    if "DATA ENTRY" in wb.sheetnames:
        _import_entries(con, wb["DATA ENTRY"], report)
    if "REWORK EFFICIENCY" in wb.sheetnames:
        _import_rework(con, wb["REWORK EFFICIENCY"], report)
    con.commit()
    return report


def _import_masters(con, ws, report):
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=4).value
        if name:
            cur = con.execute(
                """INSERT OR IGNORE INTO parts
                   (customer,project,name,part_no,cycle_time_s,total_cavity,running_cavity,
                    shot_wt_g,part_wt_g,runner_wt_g,material_grade)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value,
                 str(name).strip(), ws.cell(row=r, column=5).value,
                 _n(ws.cell(row=r, column=6).value), _n(ws.cell(row=r, column=7).value),
                 _n(ws.cell(row=r, column=8).value), _n(ws.cell(row=r, column=9).value),
                 _n(ws.cell(row=r, column=10).value), _n(ws.cell(row=r, column=11).value),
                 ws.cell(row=r, column=12).value))
            report["parts_added"] += cur.rowcount
        for col, table in ((14, "supervisors"), (16, "operators"), (18, "machines")):
            v = ws.cell(row=r, column=col).value
            if v and not str(v).startswith("↑"):
                cur = con.execute(f"INSERT OR IGNORE INTO {table}(name) VALUES(?)",
                                  (str(v).strip(),))
                report["masters_added"] += cur.rowcount


def _import_entries(con, ws, report):
    shift_minutes = float(get_setting(con, "shift_minutes", "720"))
    for row in ws.iter_rows(min_row=5, values_only=True):
        date = _cell_date(row[COL_DATE - 1])
        if not date:
            continue
        machine = row[COL_MACHINE - 1]
        part = row[COL_PART - 1]
        production = _n(row[COL_PRODUCTION - 1])
        if not (machine and part):
            report["skipped"].append(f"{date}: missing machine/part")
            continue
        part = str(part).strip()
        part_row = con.execute("SELECT * FROM parts WHERE name=?", (part,)).fetchone()
        if part_row is None:
            report["skipped"].append(f"{date} {machine}: unknown part '{part}'")
            continue
        defects = {}
        for i, name in enumerate(DEFECT_TYPES):
            v = _n(row[DEFECT_COL_START - 1 + i])
            if v:
                defects[name] = int(v)
        downtime = {}
        for i, name in enumerate(DOWNTIME_REASONS):
            v = _n(row[DOWNTIME_COL_START - 1 + i])
            if v:
                downtime[name] = v
        planned = _n(row[COL_PLANNED - 1])
        c = calc.compute_entry(part_row, planned, production, defects, downtime, shift_minutes)

        # cross-check against the workbook's own computed cells
        for label, ours, theirs in (("OK", c["ok_qty"], _n(row[COL_OK - 1])),
                                    ("REJ", c["rejection"], _n(row[COL_REJ - 1])),
                                    ("OEE", c["oee"], _n(row[COL_OEE - 1]))):
            if ours is not None and theirs is not None and abs(ours - theirs) > 0.01:
                report["mismatches"].append(
                    f"{date} {machine} {part}: {label} recomputed {ours} vs workbook {theirs}")

        key = (date, machine, row[COL_SHIFT - 1], part, row[COL_OP - 1])
        old = con.execute(
            "SELECT id FROM shift_entries WHERE date=? AND machine=? AND shift=? AND part=? "
            "AND operator IS ?", key).fetchall()
        for o in old:
            con.execute("DELETE FROM shift_entries WHERE id=?", (o["id"],))
            report["replaced"] += 1
        cur = con.execute(
            """INSERT INTO shift_entries
               (date,machine,shift,supervisor,operator,part,material,cycle_time_s,
                planned_min,production,remarks,
                total_dt,avail_time,eff_time,tgt_hr,tgt_qty,ok_qty,rejection,
                eff_pct,rej_pct,ok_wt_kg,rej_wt_kg,runner_kg,total_rm_kg,
                availability,performance,quality,oee)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (date, machine, row[COL_SHIFT - 1], row[COL_SUP - 1], row[COL_OP - 1], part,
             part_row["material_grade"], part_row["cycle_time_s"],
             planned, int(production) if production is not None else None,
             row[COL_REMARKS - 1],
             c["total_dt"], c["avail_time"], c["eff_time"], c["tgt_hr"], c["tgt_qty"],
             c["ok_qty"], c["rejection"], c["eff_pct"], c["rej_pct"],
             c["ok_wt_kg"], c["rej_wt_kg"], c["runner_kg"], c["total_rm_kg"],
             c["availability"], c["performance"], c["quality"], c["oee"]))
        entry_id = cur.lastrowid
        for name, qty in defects.items():
            con.execute("INSERT INTO entry_defects(entry_id,defect,qty) VALUES(?,?,?)",
                        (entry_id, name, qty))
        for name, minutes in downtime.items():
            con.execute("INSERT INTO entry_downtime(entry_id,reason,minutes) VALUES(?,?,?)",
                        (entry_id, name, minutes))
        report["entries"] += 1


def _import_rework(con, ws, report):
    for row in ws.iter_rows(min_row=6, max_col=10, values_only=True):
        date = _cell_date(row[0])
        if not (date and row[1]):
            continue
        exists = con.execute(
            "SELECT 1 FROM rework_entries WHERE date=? AND operator=? AND description IS ?",
            (date, row[1], row[4])).fetchone()
        if exists:
            continue
        con.execute(
            """INSERT INTO rework_entries
               (date,operator,shift,work_type,description,target_qty,ok_qty,rej_qty,remarks)
               VALUES(?,?,?,?,?,?,?,?,?)""",
            (date, row[1], row[2], row[3], row[4],
             _n(row[5]), _n(row[6]), _n(row[7]), row[9]))
        report["rework"] += 1
